#Einlesen Module
import pandas as pd
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog


#Allgemeine Funktion definieren
def read_csv_data(source_file, object_type, quantity_column):
    
    #CSV File einlesen
    base_df = pd.read_csv(source_file,sep=";")

    #Eingelesene CSV zeigen
    print("Ergebnis eingelesenes CSV: {}".format(str(base_df)))

    #Filtern der Liste nach Objekten
    filtered_data = base_df[base_df["Objekt"] == object_type]

    #Summe aus den gefilterten Daten
    current_sum = filtered_data[quantity_column].sum()

    #Ergebnis Filter zeigen
    print("Ergebnis Objektfilter: {}".format(str(filtered_data)))

    return current_sum


#GUI
def choose_folder_and_load_csv():
    folder_path = filedialog.askdirectory()
    csv_files = [file for file in os.listdir(folder_path) if file.endswith(".csv")]

    if csv_files:
        selected_csv_file = os.path.join(folder_path, csv_files[0])
        print("Ausgewählte CSV-Datei:", selected_csv_file)

        df = pd.read_csv(selected_csv_file)
        print("Inhalt der CSV-Datei:")
        print(df.head())

    return folder_path

root = tk.Tk()
root.title("Import CSV-Datei")

label = tk.Label(root, text="Wähle den Ordner, in dem die CSV-Dateien abgelegt sind")
label.pack(padx=10, pady=20)

button = tk.Button(root, text="Auswählen", command=choose_folder_and_load_csv)
button.pack(pady=10)

root.mainloop()

selected_folder = choose_folder_and_load_csv()
if selected_folder:

    #Variablen für die Summen
    sum_of_quantity_walls_formwork = 0.0
    sum_of_quantity_columns_formwork = 0.0
    sum_of_quantity_ceiling_formwork = 0.0
    sum_of_quantity_parapet_formwork = 0.0

    sum_of_quantity_brick_area = 0.0
    sum_of_quantity_sandlimebrick_area = 0.0

    sum_of_quantity_floor_volume = 0.0
    sum_of_quantity_walls_volume = 0.0
    sum_of_quantity_columns_volume = 0.0
    sum_of_quantity_ceiling_volume = 0.0
    sum_of_quantity_parapet_volume = 0.0

    for file in os.listdir(selected_folder):
        if file.endswith(".csv"):
            current_csv_file = os.path.join(selected_folder, file)
            print(current_csv_file)

        #Kombinationen definieren
        sum_of_storey = read_csv_data(current_csv_file, "Wände", "Fläche Wandschalung")
        sum_of_quantity_walls_formwork += sum_of_storey

        sum_of_storey = read_csv_data(current_csv_file, "Stützen", "Wandfläche")
        sum_of_quantity_columns_formwork += sum_of_storey

        sum_of_storey = read_csv_data(current_csv_file, "Decken", "Fläche")
        sum_of_quantity_ceiling_formwork += sum_of_storey

        sum_of_storey = read_csv_data(current_csv_file, "Brüstung", "Fläche Wandschalung")
        sum_of_quantity_parapet_formwork += sum_of_storey


        sum_of_storey = read_csv_data(current_csv_file, "Backstein", "Wandfläche")
        sum_of_quantity_brick_area += sum_of_storey

        sum_of_storey = read_csv_data(current_csv_file, "Kalksandstein", "Wandfläche")
        sum_of_quantity_sandlimebrick_area += sum_of_storey


        sum_of_storey = read_csv_data(current_csv_file, "Boden", "Volumen")
        sum_of_quantity_floor_volume += sum_of_storey

        sum_of_storey = read_csv_data(current_csv_file, "Wände", "Volumen Wand")
        sum_of_quantity_walls_volume += sum_of_storey

        sum_of_storey = read_csv_data(current_csv_file, "Stützen", "Volumen Wand")
        sum_of_quantity_columns_volume += sum_of_storey

        sum_of_storey = read_csv_data(current_csv_file, "Decken", "Volumen")
        sum_of_quantity_ceiling_volume += sum_of_storey

        sum_of_storey = read_csv_data(current_csv_file, "Brüstung", "Volumen Wand")
        sum_of_quantity_parapet_volume += sum_of_storey

    
    #Excel-Datei vorbereiten
    excel_output_path = "/Users/leameier/Desktop/Programmierung 1/0_Mein Projekt/Pläne Test/Mengenermittlung OASE mit LV Vergleich.xlsx"

    workbook = openpyxl.load_workbook(excel_output_path)
    sheet_name = 'Auswertung'
    sheet = workbook[sheet_name]


    #Ergebnisse in Excel exportieren
    sheet['H7'] = sum_of_quantity_walls_formwork
    sheet['H8'] = sum_of_quantity_columns_formwork
    sheet['H9'] = sum_of_quantity_ceiling_formwork
    sheet['H10'] = sum_of_quantity_parapet_formwork

    sheet['H20'] = sum_of_quantity_brick_area
    sheet['H21'] = sum_of_quantity_sandlimebrick_area

    sheet['H12'] = sum_of_quantity_floor_volume
    sheet['H13'] = sum_of_quantity_walls_volume
    sheet['H14'] = sum_of_quantity_columns_volume
    sheet['H15'] = sum_of_quantity_ceiling_volume
    sheet['H16'] = sum_of_quantity_parapet_volume

    workbook.save(excel_output_path)
    workbook.close()


    #Abschlussmeldung
    print(f'Die Ergebnisse wurden erfolgreich in die Excel-Datei exportiert: {excel_output_path}')

else:
    print("Kein Ordner ausgewählt")








###############################

##Iteration durch die Dateien
#source_folder = '/Users/leameier/Desktop/Programmierung 1/0_Mein Projekt/Vorlagen/Pläne Ausmass'

##Variablen für die Summen
#sum_of_quantity_walls_formwork = 0.0
#sum_of_quantity_columns_formwork = 0.0
#sum_of_quantity_ceiling_formwork = 0.0
#sum_of_quantity_parapet_formwork = 0.0

#sum_of_quantity_brick_area = 0.0
#sum_of_quantity_sandlimebrick_area = 0.0

#sum_of_quantity_floor_volume = 0.0
#sum_of_quantity_walls_volume = 0.0
#sum_of_quantity_columns_volume = 0.0
#sum_of_quantity_ceiling_volume = 0.0
#sum_of_quantity_parapet_volume = 0.0

#for file in os.listdir(source_folder):
    
    #if file.endswith(".csv"):
        #current_csv_file = os.path.join(source_folder, file)
        #print(current_csv_file)

        ##Kombinationen definieren
        #sum_of_storey = read_csv_data(current_csv_file, "Wände", "Fläche Wandschalung")
        #sum_of_quantity_walls_formwork += sum_of_storey

        #sum_of_storey = read_csv_data(current_csv_file, "Stützen", "Wandfläche")
        #sum_of_quantity_columns_formwork += sum_of_storey

        #sum_of_storey = read_csv_data(current_csv_file, "Decken", "Fläche")
        #sum_of_quantity_ceiling_formwork += sum_of_storey

        #sum_of_storey = read_csv_data(current_csv_file, "Brüstung", "Fläche Wandschalung")
        #sum_of_quantity_parapet_formwork += sum_of_storey


        #sum_of_storey = read_csv_data(current_csv_file, "Backstein", "Wandfläche")
        #sum_of_quantity_brick_area += sum_of_storey

        #sum_of_storey = read_csv_data(current_csv_file, "Kalksandstein", "Wandfläche")
        #sum_of_quantity_sandlimebrick_area += sum_of_storey


        #sum_of_storey = read_csv_data(current_csv_file, "Boden", "Volumen")
        #sum_of_quantity_floor_volume += sum_of_storey

        #sum_of_storey = read_csv_data(current_csv_file, "Wände", "Volumen Wand")
        #sum_of_quantity_walls_volume += sum_of_storey

        #sum_of_storey = read_csv_data(current_csv_file, "Stützen", "Volumen Wand")
        #sum_of_quantity_columns_volume += sum_of_storey

        #sum_of_storey = read_csv_data(current_csv_file, "Decken", "Volumen")
        #sum_of_quantity_ceiling_volume += sum_of_storey

        #sum_of_storey = read_csv_data(current_csv_file, "Brüstung", "Volumen Wand")
        #sum_of_quantity_parapet_volume += sum_of_storey

    
##Excel-Datei vorbereiten
#excel_output_path = "/Users/leameier/Desktop/Programmierung 1/0_Mein Projekt/Vorlagen/Pläne Ausmass/Mengenermittlung OASE mit LV Vergleich.xlsx"

#workbook = openpyxl.load_workbook(excel_output_path)
#sheet_name = 'Auswertung'
#sheet = workbook[sheet_name]


##Ergebnisse in Excel exportieren
#sheet['H7'] = sum_of_quantity_walls_formwork
##sheet['H8'] = sum_of_quantity_columns_formwork
#sheet['H9'] = sum_of_quantity_ceiling_formwork
#sheet['H10'] = sum_of_quantity_parapet_formwork

#sheet['H20'] = sum_of_quantity_brick_area
##sheet['H21'] = sum_of_quantity_sandlimebrick_area

#sheet['H12'] = sum_of_quantity_floor_volume
#sheet['H13'] = sum_of_quantity_walls_volume
#sheet['H14'] = sum_of_quantity_columns_volume
#sheet['H15'] = sum_of_quantity_ceiling_volume
#sheet['H16'] = sum_of_quantity_parapet_volume

#workbook.save(excel_output_path)
#workbook.close()


##Abschlussmeldung
#print(f'Die Ergebnisse wurden erfolgreich in die Excel-Datei exportiert: {excel_output_path}')








############################################

##Pfad bestimmen von File
#base_file_path_1UG_ver = "/Users/leameier/Desktop/Programmierung 1/0_Mein Projekt/Vorlagen/CSV neu/Wände/02_Grundriss 1.UG_ver.csv"
#base_file_path_0EG_ver = "/Users/leameier/Desktop/Programmierung 1/0_Mein Projekt/Vorlagen/CSV neu/Wände/03_Grundriss 0.EG_ver.csv"
#base_file_path_1OG_ver = "/Users/leameier/Desktop/Programmierung 1/0_Mein Projekt/Vorlagen/CSV neu/Wände/04_Grundriss 1.OG_ver.csv"


#base_file_path_2UG_hor = "/Users/leameier/Desktop/Programmierung 1/0_Mein Projekt/Vorlagen/CSV neu/Decken/01_Grundriss 2.UG_hor.csv"
#base_file_path_1UG_hor = "/Users/leameier/Desktop/Programmierung 1/0_Mein Projekt/Vorlagen/CSV neu/Decken/02_Grundriss 1.UG_hor.csv"


##CSV File einlesen
#base_df_1UG_ver = pd.read_csv(base_file_path_1UG_ver,sep=";")
#base_df_0EG_ver = pd.read_csv(base_file_path_0EG_ver,sep=";")
#base_df_1OG_ver = pd.read_csv(base_file_path_1OG_ver,sep=";")


#base_df_2UG_hor = pd.read_csv(base_file_path_2UG_hor,sep=";")
#base_df_1UG_hor = pd.read_csv(base_file_path_1UG_hor,sep=";")

##Eingelesene CSV zeigen
#print("Ergebnis eingelesenes CSV 1UG_ver: {}".format(str(base_df_1UG_ver)))
#print("Ergebnis eingelesenes CSV 0EG_ver: {}".format(str(base_df_0EG_ver)))
#print("Ergebnis eingelesenes CSV 1OG_ver: {}".format(str(base_df_1OG_ver)))


#print("Ergebnis eingelesenes CSV 2UG_hor: {}".format(str(base_df_2UG_hor)))
#print("Ergebnis eingelesenes CSV 1UG_hor: {}".format(str(base_df_1UG_hor)))


##Filtern der Liste nach Objekten
#wände_df_1UG_ver = base_df_1UG_ver[base_df_1UG_ver["Objekt"] == "Wände"]
#stützen_df_1UG_ver = base_df_1UG_ver[base_df_1UG_ver["Objekt"] == "Stützen"]
#brüstung_df_1UG_ver = base_df_1UG_ver[base_df_1UG_ver["Objekt"] == "Brüstung"]
#backstein_df_1UG_ver = base_df_1UG_ver[base_df_1UG_ver["Objekt"] == "Backstein"]
#kalksandstein_df_1UG_ver = base_df_1UG_ver[base_df_1UG_ver["Objekt"] == "Kalksandstein"]

#ände_df_0EG_ver = base_df_0EG_ver[base_df_0EG_ver["Objekt"] == "Wände"]
#stützen_df_0EG_ver = base_df_0EG_ver[base_df_0EG_ver["Objekt"] == "Stützen"]
#brüstung_df_0EG_ver = base_df_0EG_ver[base_df_0EG_ver["Objekt"] == "Brüstung"]
#backstein_df_0EG_ver = base_df_0EG_ver[base_df_0EG_ver["Objekt"] == "Backstein"]
#kalksandstein_df_0EG_ver = base_df_0EG_ver[base_df_0EG_ver["Objekt"] == "Kalksandstein"]

#wände_df_1OG_ver = base_df_1OG_ver[base_df_1OG_ver["Objekt"] == "Wände"]
#stützen_df_1OG_ver = base_df_1OG_ver[base_df_1OG_ver["Objekt"] == "Stützen"]
#brüstung_df_1OG_ver = base_df_1OG_ver[base_df_1OG_ver["Objekt"] == "Brüstung"]
#backstein_df_1OG_ver = base_df_1OG_ver[base_df_1OG_ver["Objekt"] == "Backstein"]
#kalksandstein_df_1OG_ver = base_df_1OG_ver[base_df_1OG_ver["Objekt"] == "Kalksandstein"]


#boden_df_2UG_hor = base_df_2UG_hor[base_df_2UG_hor["Objekt"] == "Boden"]
#decken_df_2UG_hor = base_df_2UG_hor[base_df_2UG_hor["Objekt"] == "Decken"]

#boden_df_1UG_hor = base_df_1UG_hor[base_df_1UG_hor["Objekt"] == "Boden"]
#decken_df_1UG_hor = base_df_1UG_hor[base_df_1UG_hor["Objekt"] == "Decken"]


##Ergebnis Filter zeigen
#print("Ergebnis Objektfilter Wände 1UG_ver: {}".format(str(wände_df_1UG_ver)))
#print("Ergebnis Objektfilter Stützen 1UG_ver: {}".format(str(stützen_df_1UG_ver)))
#print("Ergebnis Objektfilter Brüstung 1UG_ver: {}".format(str(brüstung_df_1UG_ver)))
#print("Ergebnis Objektfilter Backstein 1UG_ver: {}".format(str(backstein_df_1UG_ver)))
#print("Ergebnis Objektfilter Kalksandstein 1UG_ver: {}".format(str(kalksandstein_df_1UG_ver)))

#print("Ergebnis Objektfilter Wände 0EG_ver: {}".format(str(wände_df_0EG_ver)))
#print("Ergebnis Objektfilter Stützen 0EG_ver: {}".format(str(stützen_df_0EG_ver)))
#print("Ergebnis Objektfilter Brüstung 0EG_ver: {}".format(str(brüstung_df_0EG_ver)))
#print("Ergebnis Objektfilter Backstein 0EG_ver: {}".format(str(backstein_df_0EG_ver)))
#print("Ergebnis Objektfilter Kalksandstein 0EG_ver: {}".format(str(kalksandstein_df_0EG_ver)))

#print("Ergebnis Objektfilter Wände 1OG_ver: {}".format(str(wände_df_1OG_ver)))
#print("Ergebnis Objektfilter Stützen 1OG_ver: {}".format(str(stützen_df_1OG_ver)))
#print("Ergebnis Objektfilter Brüstung 1OG_ver: {}".format(str(brüstung_df_1OG_ver)))
#print("Ergebnis Objektfilter Backstein 1OG_ver: {}".format(str(backstein_df_1OG_ver)))
#print("Ergebnis Objektfilter Kalksandstein 1OG_ver: {}".format(str(kalksandstein_df_1OG_ver)))


#print("Ergebnis Objektfilter Boden 2UG_hor: {}".format(str(boden_df_2UG_hor)))
#print("Ergebnis Objektfilter Decken 2UG_hor: {}".format(str(decken_df_2UG_hor)))

#print("Ergebnis Objektfilter Boden 1UG_hor: {}".format(str(boden_df_1UG_hor)))
#print("Ergebnis Objektfilter Decken 1UG_hor: {}".format(str(decken_df_1UG_hor)))

##Summe aus den gefilterten Daten (Fläche)
#Fläche_Wände_1UG_ver = wände_df_1UG_ver["Fläche Wandschalung"].sum()
#Fläche_Stützen_1UG_ver = stützen_df_1UG_ver["Wandfläche"].sum()
#Fläche_Brüstung_1UG_ver = brüstung_df_1UG_ver["Fläche Wandschalung"].sum()
#Fläche_Backstein_1UG_ver = backstein_df_1UG_ver["Wandfläche"].sum()
#Fläche_Kalksandstein_1UG_ver = kalksandstein_df_1UG_ver["Wandfläche"].sum()

#Fläche_Wände_0EG_ver = wände_df_0EG_ver["Fläche Wandschalung"].sum()
#Fläche_Stützen_0EG_ver = stützen_df_0EG_ver["Wandfläche"].sum()
#Fläche_Brüstung_0EG_ver = brüstung_df_0EG_ver["Fläche Wandschalung"].sum()
#Fläche_Backstein_0EG_ver = backstein_df_0EG_ver["Wandfläche"].sum()
#Fläche_Kalksandstein_0EG_ver = kalksandstein_df_0EG_ver["Wandfläche"].sum()

#Fläche_Wände_1OG_ver = wände_df_1OG_ver["Fläche Wandschalung"].sum()
#Fläche_Stützen_1OG_ver = stützen_df_1OG_ver["Wandfläche"].sum()
#Fläche_Brüstung_1OG_ver = brüstung_df_1OG_ver["Fläche Wandschalung"].sum()
#Fläche_Backstein_1OG_ver = backstein_df_1OG_ver["Wandfläche"].sum()
#Fläche_Kalksandstein_1OG_ver = kalksandstein_df_1OG_ver["Wandfläche"].sum()


#Fläche_Decken_2UG_hor = decken_df_2UG_hor["Fläche"].sum()

#Fläche_Decken_1UG_hor = decken_df_1UG_hor["Fläche"].sum()

##Summe aus den gefilterten Daten (Volumen)
#Volumen_Wände_1UG_ver = wände_df_1UG_ver["Volumen Wand"].sum()
#Volumen_Stützen_1UG_ver = stützen_df_1UG_ver["Volumen Wand"].sum()
#Volumen_Brüstung_1UG_ver = brüstung_df_1UG_ver["Volumen Wand"].sum()

#Volumen_Wände_0EG_ver = wände_df_0EG_ver["Volumen Wand"].sum()
#Volumen_Stützen_0EG_ver = stützen_df_0EG_ver["Volumen Wand"].sum()
#Volumen_Brüstung_0EG_ver = brüstung_df_0EG_ver["Volumen Wand"].sum()

#Volumen_Wände_1OG_ver = wände_df_1OG_ver["Volumen Wand"].sum()
#Volumen_Stützen_1OG_ver = stützen_df_1OG_ver["Volumen Wand"].sum()
#Volumen_Brüstung_1OG_ver = brüstung_df_1OG_ver["Volumen Wand"].sum()


#Volumen_Boden_2UG_hor = boden_df_2UG_hor["Volumen"].sum()
#Volumen_Decken_2UG_hor = decken_df_2UG_hor["Volumen"].sum()

#Volumen_Boden_1UG_hor = boden_df_1UG_hor["Volumen"].sum()
#Volumen_Decken_1UG_hor = decken_df_1UG_hor["Volumen"].sum()

##Ergebnis Summe zeigen (Fläche)
#print("Summe Schalung Wände 1UG_ver: {}".format(str(Fläche_Wände_1UG_ver)))
#print("Summe Fläche Stützen 1UG_ver: {}".format(str(Fläche_Stützen_1UG_ver)))
#print("Summe Schalung Brüstung 1UG_ver: {}".format(str(Fläche_Brüstung_1UG_ver)))
#print("Summe Fläche Backstein 1UG_ver: {}".format(str(Fläche_Backstein_1UG_ver)))
#print("Summe Fläche Kalksandstein 1UG_ver: {}".format(str(Fläche_Kalksandstein_1UG_ver)))

#print("Summe Schalung Wände 0EG_ver: {}".format(str(Fläche_Wände_0EG_ver)))
#print("Summe Fläche Stützen 0EG_ver: {}".format(str(Fläche_Stützen_0EG_ver)))
#print("Summe Schalung Brüstung 0EG_ver: {}".format(str(Fläche_Brüstung_0EG_ver)))
#print("Summe Fläche Backstein 0EG_ver: {}".format(str(Fläche_Backstein_0EG_ver)))
#print("Summe Fläche Kalksandstein 0EG_ver: {}".format(str(Fläche_Kalksandstein_0EG_ver)))

#print("Summe Schalung Wände 1OG_ver: {}".format(str(Fläche_Wände_1OG_ver)))
#print("Summe Fläche Stützen 1OG_ver: {}".format(str(Fläche_Stützen_1OG_ver)))
#print("Summe Schalung Brüstung 1OG_ver: {}".format(str(Fläche_Brüstung_1OG_ver)))
#print("Summe Fläche Backstein 1OG_ver: {}".format(str(Fläche_Backstein_1OG_ver)))
#print("Summe Fläche Kalksandstein 1OG_ver: {}".format(str(Fläche_Kalksandstein_1OG_ver)))


#print("Summe Fläche Decken 2UG_hor: {}".format(str(Fläche_Decken_2UG_hor)))

#print("Summe Fläche Decken 1UG_hor: {}".format(str(Fläche_Decken_1UG_hor)))

##Ergebnis Summe zeigen (Volumen)
#print("Summe Volumen Wände 1UG_ver: {}".format(str(Volumen_Wände_1UG_ver)))
#print("Summe Volumen Stützen 1UG_ver: {}".format(str(Volumen_Stützen_1UG_ver)))
#print("Summe Volumen Brüstung 1UG_ver: {}".format(str(Volumen_Brüstung_1UG_ver)))

#print("Summe Volumen Wände 0EG_ver: {}".format(str(Volumen_Wände_0EG_ver)))
#print("Summe Volumen Stützen 0EG_ver: {}".format(str(Volumen_Stützen_0EG_ver)))
#print("Summe Volumen Brüstung 0EG_ver: {}".format(str(Volumen_Brüstung_0EG_ver)))

#print("Summe Volumen Wände 1OG_ver: {}".format(str(Volumen_Wände_1OG_ver)))
#print("Summe Volumen Stützen 1OG_ver: {}".format(str(Volumen_Stützen_1OG_ver)))
#print("Summe Volumen Brüstung 1OG_ver: {}".format(str(Volumen_Brüstung_1OG_ver)))

#print("Summe Volumen Boden 2UG_hor: {}".format(str(Volumen_Boden_2UG_hor)))
#print("Summe Volumen Decken 2UG_hor: {}".format(str(Volumen_Decken_2UG_hor)))

#print("Summe Volumen Boden 1UG_hor: {}".format(str(Volumen_Boden_1UG_hor)))
#print("Summe Volumen Decken 1UG_hor: {}".format(str(Volumen_Decken_1UG_hor)))


##Gesamtsumme aus allen Files und gefilterten Daten (Fläche)
#Wände_summe_gesamt_Fläche_ver = Fläche_Wände_1UG_ver + Fläche_Wände_0EG_ver + Fläche_Wände_1OG_ver
#Stützen_summe_gesamt_Fläche_ver = Fläche_Stützen_1UG_ver + Fläche_Stützen_0EG_ver + Fläche_Stützen_1OG_ver
#Brüstung_summe_gesamt_Fläche_ver = Fläche_Brüstung_1UG_ver + Fläche_Brüstung_0EG_ver + Fläche_Brüstung_1OG_ver
#Backstein_summe_gesamt_Fläche_ver = Fläche_Backstein_1UG_ver + Fläche_Backstein_0EG_ver + Fläche_Backstein_1OG_ver
#Kalksandstein_summe_gesamt_Fläche_ver = Fläche_Kalksandstein_1UG_ver + Fläche_Kalksandstein_0EG_ver + Fläche_Kalksandstein_1OG_ver


#Decken_summe_gesamt_Fläche_hor = Fläche_Decken_1UG_hor + Fläche_Decken_2UG_hor

##Gesamtsumme aus allen Files und gefilterten Daten (Volumen)
#Wände_summe_gesamt_Volumen_ver = Volumen_Wände_1UG_ver + Volumen_Wände_0EG_ver + Volumen_Wände_1OG_ver
#Stützen_summe_gesamt_Volumen_ver = Volumen_Stützen_1UG_ver + Volumen_Stützen_0EG_ver + Volumen_Stützen_1OG_ver
#Brüstung_summe_gesamt_Volumen_ver = Volumen_Brüstung_1UG_ver + Volumen_Brüstung_0EG_ver + Volumen_Brüstung_1OG_ver


#Boden_summe_gesamt_Volumen_hor = Volumen_Boden_1UG_hor + Volumen_Boden_2UG_hor
#Decken_summe_gesamt_Volumen_hor = Volumen_Decken_1UG_hor + Volumen_Decken_2UG_hor


##Ergebnis Gesamtsumme zeigen (Fläche)
#print("Gesamtsumme Schalung Wände_ver: {}".format(str(Wände_summe_gesamt_Fläche_ver)))
#print("Gesamtsumme Fläche Stützen_ver: {}".format(str(Stützen_summe_gesamt_Fläche_ver)))
#print("Gesamtsumme Schalung Brüstung_ver: {}".format(str(Brüstung_summe_gesamt_Fläche_ver)))
#print("Gesamtsumme Fläche Backstein_ver: {}".format(str(Backstein_summe_gesamt_Fläche_ver)))
#print("Gesamtsumme Fläche Kalksandstein_ver: {}".format(str(Kalksandstein_summe_gesamt_Fläche_ver)))


#print("Gesamtsumme Fläche Decken_hor: {}".format(str(Decken_summe_gesamt_Fläche_hor)))

##Ergebnis Gesamtsumme zeigen (Volumen)
#print("Gesamtsumme Volumen Wände_ver: {}".format(str(Wände_summe_gesamt_Volumen_ver)))
#print("Gesamtsumme Volumen Stützen_ver: {}".format(str(Stützen_summe_gesamt_Volumen_ver)))
#print("Gesamtsumme Volumen Brüstung_ver: {}".format(str(Brüstung_summe_gesamt_Volumen_ver)))


#print("Gesamtsumme Volumen Boden_hor: {}".format(str(Boden_summe_gesamt_Volumen_hor)))
#print("Gesamtsumme Volumen Decken_hor: {}".format(str(Decken_summe_gesamt_Volumen_hor)))


##Excel-Datei vorbereiten
#excel_output_path = "/Users/leameier/Desktop/Programmierung 1/0_Mein Projekt/Vorlagen/Mengenermittlung OASE mit LV Vergleich.xlsx"

#workbook = openpyxl.load_workbook(excel_output_path)
#sheet_name = 'Tabelle1'
#sheet = workbook[sheet_name]


##Ergebnisse in Excel exportieren
#sheet['H7'] = Wände_summe_gesamt_Fläche_ver
#sheet['H8'] = Stützen_summe_gesamt_Fläche_ver
#heet['H10'] = Brüstung_summe_gesamt_Fläche_ver
#sheet['H20'] = Backstein_summe_gesamt_Fläche_ver 
#sheet['H21'] = Kalksandstein_summe_gesamt_Fläche_ver 

#sheet['H9'] = Decken_summe_gesamt_Fläche_hor

#sheet['H13'] = Wände_summe_gesamt_Volumen_ver
#sheet['H14'] = Stützen_summe_gesamt_Volumen_ver
#sheet['H16'] = Brüstung_summe_gesamt_Volumen_ver

#sheet['H12'] = Boden_summe_gesamt_Volumen_hor
#sheet['H15'] = Decken_summe_gesamt_Volumen_hor


##Ergebnisse in Excel speichern
#workbook.save(excel_output_path)
#workbook.close()


##Abschlussmeldung
#print(f'Die Ergebnisse wurden erfolgreich in die Excel-Datei exportiert: {excel_output_path}')